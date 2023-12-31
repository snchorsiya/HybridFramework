import openpyxl
from openpyxl.styles import PatternFill


def get_row_count(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_row


def get_column_count(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_column


def read_data(file, sheet_name, row_name, column_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.cell(row_name, column_name).value


def write_data(file, sheet_name, row_name, column_name, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    sheet.cell(row_name, column_name).value = data
    workbook.save(file)


def fill_green(file, sheet_name, row_name, column_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', patternType='solid')
    sheet.cell(row_name, column_name).fill = green_fill
    workbook.save(file)


def fill_red(file, sheet_name, row_name, column_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', patternType='solid')
    sheet.cell(row_name, column_name).fill = red_fill
    workbook.save(file)




